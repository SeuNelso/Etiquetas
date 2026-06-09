from __future__ import annotations

import os
import sys
import win32print
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from datetime import datetime
from typing import Literal, Optional

CodigoEtiqueta = Literal["barcode", "qrcode"]

from openpyxl import load_workbook

# CONFIG
DEFAULT_PRINTER = "ZDesigner GK420d"
_BASE = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(_BASE, "historico.txt")
CONFIG_FILE = os.path.join(_BASE, "Config.txt")
_PRINTER_PREFS_LEGACY = os.path.join(_BASE, "impressora_preferida.txt")
_CODIGO_PREFS_LEGACY = os.path.join(_BASE, "codigo_etiqueta_preferido.txt")
# Coloque artigos.xlsx na mesma pasta que o .exe (ou este .py).
ARTIGOS_XLSX = os.path.join(_BASE, "artigos.xlsx")

# Primeira linha pode ser cabeçalho (Codigo / Descricao) ou já ser dados.
_HEADERS = frozenset({"codigo", "código", "code", "ref", "artigo", "sku", "descrição", "descricao", "description", "desc"})


def _normalizar_chave(codigo: str) -> str:
    return codigo.strip().casefold()


def _cel_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def carregar_artigos(path: str) -> dict:
    """Devolve mapa codigo_normalizado -> descricao."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"Ficheiro não encontrado:\n{path}") from None
    except Exception as e:
        raise RuntimeError(f"Erro ao abrir o Excel:\n{e}") from e

    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return {}

    start = 0
    a0 = _cel_str(rows[0][0]).lower() if rows[0] else ""
    b0 = _cel_str(rows[0][1]).lower() if rows[0] and len(rows[0]) > 1 else ""
    if a0 in _HEADERS or b0 in _HEADERS:
        start = 1

    out = {}
    for row in rows[start:]:
        if not row:
            continue
        codigo = _cel_str(row[0])
        if not codigo:
            continue
        desc = _cel_str(row[1]) if len(row) > 1 else ""
        k = _normalizar_chave(codigo)
        if k not in out:
            out[k] = desc
    return out


def _zpl_safe(text: str) -> str:
    return " ".join(text.replace("^", " ").replace("~", " ").split())


def listar_impressoras() -> list[str]:
    """Nomes de impressoras instaladas / ligadas ao utilizador (inclui rede)."""
    flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
    try:
        raw = win32print.EnumPrinters(flags, None, 1)
    except Exception:
        return []
    seen: set[str] = set()
    out: list[str] = []
    for row in raw:
        name = row[2] if len(row) > 2 else ""
        if not name or name in seen:
            continue
        seen.add(name)
        out.append(name)
    out.sort(key=str.casefold)
    return out


def _parse_config(text: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, _, v = line.partition("=")
            out[k.strip().lower()] = v.strip()
    return out


def _format_config(data: dict[str, str]) -> str:
    impressora = data.get("impressora", "")
    codigo = data.get("codigo", "barcode").lower()
    if codigo != "qrcode":
        codigo = "barcode"
    return f"impressora={impressora}\ncodigo={codigo}\n"


def _ler_config_ficheiro() -> dict[str, str]:
    try:
        with open(CONFIG_FILE, encoding="utf-8") as f:
            return _parse_config(f.read())
    except OSError:
        return {}


def _migrar_prefs_legado() -> dict[str, str]:
    data: dict[str, str] = {}
    try:
        with open(_PRINTER_PREFS_LEGACY, encoding="utf-8") as f:
            s = f.read().strip()
            if s:
                data["impressora"] = s
    except OSError:
        pass
    try:
        with open(_CODIGO_PREFS_LEGACY, encoding="utf-8") as f:
            s = f.read().strip().lower()
            if s == "qrcode":
                data["codigo"] = "qrcode"
            elif s:
                data["codigo"] = "barcode"
    except OSError:
        pass
    return data


def _escrever_config(data: dict[str, str]) -> None:
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(_format_config(data))
    except OSError:
        pass


def _carregar_config() -> dict[str, str]:
    if os.path.isfile(CONFIG_FILE):
        return _ler_config_ficheiro()
    legacy = _migrar_prefs_legado()
    if legacy:
        _escrever_config(legacy)
    return legacy


def _guardar_config(
    *,
    impressora: Optional[str] = None,
    codigo: Optional[CodigoEtiqueta] = None,
) -> None:
    data = _carregar_config()
    if impressora is not None:
        data["impressora"] = impressora
    if codigo is not None:
        data["codigo"] = codigo
    _escrever_config(data)


def _carregar_impressora_preferida() -> Optional[str]:
    s = _carregar_config().get("impressora", "").strip()
    return s or None


def _guardar_impressora_preferida(name: str) -> None:
    _guardar_config(impressora=name)


def _carregar_codigo_preferido() -> CodigoEtiqueta:
    if _carregar_config().get("codigo", "").lower() == "qrcode":
        return "qrcode"
    return "barcode"


def _guardar_codigo_preferido(tipo: CodigoEtiqueta) -> None:
    _guardar_config(codigo=tipo)


def gerar_zpl(
    sn: str,
    codigo_artigo: str,
    descricao: str,
    codigo_tipo: CodigoEtiqueta = "barcode",
) -> str:
    desc = _zpl_safe(descricao)
    cod = _zpl_safe(codigo_artigo)
    sn_s = _zpl_safe(sn)
    if codigo_tipo == "qrcode":
        codigo_sn = f"^FO200,30^BQN,2,4^FDMA,{sn_s}^FS"
        y_desc, y_cod = 220, 265
        ll = 300
    else:
        codigo_sn = f"^FO40,95^BY2,3,72^BCN,72,Y,N,N^FD{sn_s}^FS"
        y_desc, y_cod = 220, 265
        ll = 300
    return f"""
^XA
^PW560
^LL{ll}

^FO20,20^XGLOGO.GRF,1,1^FS

^FO40,55^A0N,30,30^FDSN:^FS
^FO140,55^A0N,30,30^FD{sn_s}^FS

{codigo_sn}

^FO40,{y_desc}^A0N,26,26^FD{desc}^FS

^FO40,{y_cod}^A0N,22,22^FD{cod}^FS

^XZ
"""


def imprimir(
    sn: str,
    codigo_artigo: str,
    descricao: str,
    printer_name: str,
    codigo_tipo: CodigoEtiqueta = "barcode",
) -> None:
    zpl = gerar_zpl(sn, codigo_artigo, descricao, codigo_tipo)

    hPrinter = win32print.OpenPrinter(printer_name)
    try:
        win32print.StartDocPrinter(hPrinter, 1, ("Etiqueta", None, "RAW"))
        try:
            win32print.StartPagePrinter(hPrinter)
            win32print.WritePrinter(hPrinter, zpl.encode("utf-8"))
            win32print.EndPagePrinter(hPrinter)
        finally:
            win32print.EndDocPrinter(hPrinter)
    finally:
        win32print.ClosePrinter(hPrinter)

    registar_log(sn, codigo_artigo, descricao)


def registar_log(sn: str, codigo_artigo: str, descricao: str) -> None:
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{datetime.now().isoformat()} | {codigo_artigo} | {descricao} | {sn}\n")


def main() -> None:
    try:
        artigos = carregar_artigos(ARTIGOS_XLSX)
    except FileNotFoundError as e:
        messagebox.showerror("Artigos", str(e))
        sys.exit(1)
    except RuntimeError as e:
        messagebox.showerror("Artigos", str(e))
        sys.exit(1)

    if not artigos:
        messagebox.showerror(
            "Artigos",
            f"O ficheiro não tem artigos válidos:\n{ARTIGOS_XLSX}\n\n"
            "Use duas colunas: código e descrição (primeira linha pode ser cabeçalho).",
        )
        sys.exit(1)

    descricao_atual: list[Optional[str]] = [None]
    codigo_atual: list[Optional[str]] = [None]

    root = tk.Tk()
    root.title("Etiquetas — Logística PT")
    root.geometry("520x470")
    root.minsize(480, 430)

    frm = tk.Frame(root, padx=12, pady=10)
    frm.pack(fill=tk.BOTH, expand=True)

    tk.Label(frm, text="Impressora", font=("Arial", 11)).pack(anchor="w")
    row_prn = tk.Frame(frm)
    row_prn.pack(fill=tk.X, pady=(0, 4))
    printer_var = tk.StringVar()
    cb_printers = ttk.Combobox(row_prn, textvariable=printer_var, state="readonly", font=("Arial", 11))
    cb_printers.pack(fill=tk.X)

    row_prn_btns = tk.Frame(frm)
    row_prn_btns.pack(fill=tk.X, pady=(0, 2))

    lbl_pref = tk.Label(frm, text="", font=("Arial", 9), fg="#555", wraplength=480, justify="left")
    lbl_pref.pack(anchor="w", pady=(0, 8))

    def refresh_pref_label() -> None:
        p = _carregar_impressora_preferida()
        lbl_pref.config(text=f"Preferida guardada: {p}" if p else "Preferida guardada: (nenhuma)")

    def escolher_impressora_inicial(nomes: list[str]) -> str:
        if not nomes:
            return ""
        pref = _carregar_impressora_preferida()
        if pref and pref in nomes:
            return pref
        if DEFAULT_PRINTER in nomes:
            return DEFAULT_PRINTER
        return nomes[0]

    def atualizar_lista_impressoras(selecionar: Optional[str] = None) -> None:
        nomes = listar_impressoras()
        cb_printers["values"] = nomes
        if not nomes:
            printer_var.set("")
            return
        atual = (selecionar or printer_var.get()).strip()
        if atual in nomes:
            printer_var.set(atual)
        else:
            printer_var.set(escolher_impressora_inicial(nomes))

    atualizar_lista_impressoras()
    refresh_pref_label()

    def on_atualizar_impressoras():
        atual = printer_var.get().strip()
        atualizar_lista_impressoras(selecionar=atual if atual else None)

    def on_definir_preferida():
        nome = printer_var.get().strip()
        if not nome:
            messagebox.showwarning(
                "Impressora preferida",
                "Selecione uma impressora na lista (ou prima «Atualizar lista»).",
            )
            return
        _guardar_impressora_preferida(nome)
        refresh_pref_label()
        messagebox.showinfo("Impressora preferida", f"A preferência foi guardada:\n{nome}")

    ttk.Button(row_prn_btns, text="Definir como preferida", command=on_definir_preferida).pack(side=tk.LEFT)
    ttk.Button(row_prn_btns, text="Atualizar lista", command=on_atualizar_impressoras).pack(side=tk.LEFT, padx=(8, 0))

    tk.Label(frm, text="Código do artigo", font=("Arial", 11)).pack(anchor="w")
    entry_artigo = tk.Entry(frm, font=("Arial", 18))
    entry_artigo.pack(fill=tk.X, pady=(0, 6))

    tk.Label(frm, text="Descrição", font=("Arial", 11)).pack(anchor="w")
    lbl_desc = tk.Label(
        frm,
        text="(introduza o código e prima Enter)",
        font=("Arial", 12),
        fg="#333",
        wraplength=480,
        justify="left",
    )
    lbl_desc.pack(anchor="w", pady=(0, 10))

    tk.Label(frm, text="Código na etiqueta", font=("Arial", 11)).pack(anchor="w")
    row_cod = tk.Frame(frm)
    row_cod.pack(fill=tk.X, pady=(0, 8))
    codigo_tipo_var = tk.StringVar(value=_carregar_codigo_preferido())
    ttk.Radiobutton(row_cod, text="Código de barras", variable=codigo_tipo_var, value="barcode").pack(
        side=tk.LEFT
    )
    ttk.Radiobutton(row_cod, text="QR Code", variable=codigo_tipo_var, value="qrcode").pack(
        side=tk.LEFT, padx=(12, 0)
    )

    tk.Label(frm, text="S/N", font=("Arial", 11)).pack(anchor="w")
    entry_sn = tk.Entry(frm, font=("Arial", 18))
    entry_sn.pack(fill=tk.X, pady=(0, 4))

    lbl_ficheiro = tk.Label(frm, text=f"Base: {os.path.basename(ARTIGOS_XLSX)}", font=("Arial", 9), fg="#666")
    lbl_ficheiro.pack(anchor="w", pady=(8, 0))

    def aplicar_artigo(event=None):
        cod = entry_artigo.get().strip()
        if not cod:
            messagebox.showwarning("Artigo", "Indique o código do artigo.")
            return "break"
        k = _normalizar_chave(cod)
        desc = artigos.get(k)
        if desc is None:
            messagebox.showerror("Artigo", f"Código não encontrado:\n{cod!r}")
            return "break"
        codigo_atual[0] = cod
        descricao_atual[0] = desc
        lbl_desc.config(text=desc, fg="#000")
        entry_sn.delete(0, tk.END)
        entry_sn.focus_set()
        return "break"

    def imprimir_sn(event=None):
        if not descricao_atual[0] or not codigo_atual[0]:
            messagebox.showwarning("Impressão", "Primeiro confirme o artigo (código + Enter).")
            return "break"
        sn = entry_sn.get().strip()
        if not sn:
            return "break"
        nome_imp = printer_var.get().strip()
        if not nome_imp:
            messagebox.showwarning(
                "Impressão",
                "Nenhuma impressora disponível.\n\n"
                "Instale uma impressora no Windows ou prima «Atualizar» depois de a ligar.",
            )
            return "break"
        tipo_cod: CodigoEtiqueta = "qrcode" if codigo_tipo_var.get() == "qrcode" else "barcode"
        try:
            imprimir(sn, codigo_atual[0], descricao_atual[0], nome_imp, tipo_cod)
            _guardar_config(impressora=nome_imp, codigo=tipo_cod)
            refresh_pref_label()
        except Exception as e:
            messagebox.showerror(
                "Erro ao imprimir",
                f"{e}\n\n"
                f"Impressora selecionada: {nome_imp!r}\n"
                "Confirme o nome em Definições > Bluetooth e dispositivos > Impressoras.",
            )
            return "break"
        entry_sn.delete(0, tk.END)
        entry_sn.focus_set()
        return "break"

    for seq in ("<Return>", "<KP_Enter>"):
        entry_artigo.bind(seq, aplicar_artigo)
        entry_sn.bind(seq, imprimir_sn)

    entry_artigo.focus_set()
    root.mainloop()


if __name__ == "__main__":
    main()
